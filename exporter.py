import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def _as_list(v):
    if isinstance(v, list):
        return v
    if isinstance(v, str) and v:
        try:
            parsed = json.loads(v)
            return parsed if isinstance(parsed, list) else [v]
        except json.JSONDecodeError:
            return [v]
    return []


def _as_dict(v):
    if isinstance(v, dict):
        return v
    if isinstance(v, str) and v:
        try:
            parsed = json.loads(v)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def _format_dms(dms) -> str:
    dms = _as_list(dms)
    parts = []
    for d in dms:
        if isinstance(d, dict):
            s = d.get("name", "")
            if d.get("role"):
                s += f" ({d['role']})"
            if d.get("contact"):
                s += f" — {d['contact']}"
            parts.append(s)
    return "; ".join(parts)


COLUMNS = [
    ("Название", 28),
    ("Домен", 24),
    ("CRM", 12),
    ("Пентест-лид", 11),
    ("Платёжеспос.", 24),
    ("Поверхность атаки", 34),
    ("Surf", 7),
    ("Данные", 10),
    ("Описание", 40),
    ("Стек", 32),
    ("Бэкенд", 16),
    ("Размер", 9),
    ("Email", 30),
    ("Телефоны", 20),
    ("Telegram", 18),
    ("WhatsApp", 16),
    ("VK", 16),
    ("ЛПР", 34),
    ("Что проверить", 40),
    ("Питч", 60),
    ("Горячесть", 10),
    ("Bug Bounty", 24),
    ("BB ссылка", 45),
    ("URL", 40),
]

_WRAP_COLS = {6, 9, 18, 19, 20}


def export_to_excel(companies: list[dict], filepath: str):
    """Export companies to a styled Excel file with all enrichment columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Компании"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, (title, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx, c in enumerate(companies, 2):
        contacts = _as_dict(c.get("contacts"))
        channels = _as_dict(c.get("channels"))
        stack = _as_list(c.get("stack"))

        row = [
            c.get("name", ""),
            c.get("domain", ""),
            c.get("crm_status", "new"),
            c.get("quality_score", ""),
            (f"{c.get('payout_score', '')}/10" + (f" · {c.get('payout_reason')}" if c.get("payout_reason") else "")),
            ", ".join(_as_list(c.get("attack_surface"))),
            c.get("surface_score", ""),
            c.get("data_sensitivity", ""),
            c.get("description", ""),
            ", ".join(stack) if stack else "",
            c.get("backend_type", ""),
            c.get("size", ""),
            ", ".join(contacts.get("emails", [])) if isinstance(contacts, dict) else "",
            ", ".join(contacts.get("phones", [])) if isinstance(contacts, dict) else "",
            ", ".join(channels.get("telegram", [])) if isinstance(channels, dict) else "",
            ", ".join(channels.get("whatsapp", [])) if isinstance(channels, dict) else "",
            ", ".join(channels.get("vk", [])) if isinstance(channels, dict) else "",
            _format_dms(c.get("decision_makers")),
            c.get("needs", ""),
            c.get("pitch", ""),
            c.get("hotness_score", ""),
            (", ".join(_as_list(c.get("bb_platforms"))) or "да") if c.get("has_bb") else "нет",
            c.get("bb_url", "") if c.get("has_bb") else "",
            c.get("url", ""),
        ]

        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=str(val) if val not in (None, "") else "")
            cell.alignment = Alignment(vertical="top", wrap_text=col in _WRAP_COLS)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
